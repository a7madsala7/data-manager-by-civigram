import pandas as pd
import openpyxl as xl
from tkinter import *
import shutil
import numpy as np
import re
import win32com.client
import datetime
from tkinter import scrolledtext
from tkinter import Tk, Label, Checkbutton, Button, Entry, Listbox, END, filedialog
import subprocess
from tkinter import messagebox
import os
# Get the current user's folder path
user_folder = os.path.expanduser("~")

# Get the name of the user's folder
user_folder_name = os.path.split(user_folder)[-1]

script_path = "script.bas"

if not os.path.isfile(script_path):
    user_folder = os.path.expanduser("~")

    # Get the name of the user's folder
    current_user = os.path.split(user_folder)[-1]
    # Define the code as a string
    code = '''Sub Main()

            ' Dimension variables
            Dim file_directory As String
            Dim file_extension As String
            Dim source_srf_file As String
            Dim SurferApp, Docs, Plot, Overlays, ContourLayer As Object

            ' Get user input for file_directory, file_extension, and source_srf_file
            file_directory = "C:\\Users\\''' + user_folder_name + '''\\Civigram" ' replace with user input
            file_extension = "txt" ' replace with user input
            source_srf_file = "C:\\Users\\''' + user_folder_name + '''\\Civigram\\sample.srf" ' replace with user input

            ' Make sure file_directory has a trailing \n
            If Right(file_directory, 1) <> "\\" Then
                file_directory = file_directory & "\\"
            End If

            'Opens Surfer
            Set SurferApp = CreateObject("Surfer.Application")
            SurferApp.Visible = False
            Set Docs = SurferApp.Documents

            'Start the loop
            file_extension = LCase(file_extension)
            grid_file = Dir(file_directory & "*." & file_extension)
            On Error GoTo FileError
            While grid_file <> ""
                'Opens a new plot document
                Set Plot = SurferApp.Documents.Add

                'Create a grid from the current grd file
                GridFile = file_directory & grid_file
                SurferApp.GridData(DataFile:=GridFile, OutGrid:=GridFile)

                'Open the source srf file
                Set Plot = Docs.Open(source_srf_file)

                'Assigns the contour map layer to the variable named "ContourLayer"
                mapnum = 1
                layernum = 1
                Set ContourLayer = Plot.Shapes.Item(mapnum).Overlays.Item(layernum)

                'Changes the grid file for the contour layer to the newly created grid file
                ContourLayer.GridFile = GridFile

                'Loops through the axes to change the tick scaling to automatic
                For i = 1 To 4
                    Plot.Shapes.Item(mapnum).Axes(i).AutoScale = True
                Next

                'Saves the SRF file to a new name
                Plot.SaveAs(FileName:=file_directory & Left(grid_file, Len(grid_file) - 4) & "_done.srf")
                Debug.Print grid_file
                Plot.Close
                grid_file = Dir() 'get next file
            Wend

            'Closes Surfer
            SurferApp.Quit
            Exit Sub

        'Error instructions
        FileError:
            Debug.Print "Error: " & grid_file & " " & Err.Description
            Resume Next
        End Sub'''

    # Replace "elkin" with the current Windows user
    code = code.replace("elkin", user_folder_name)

    # Save the code as a text file
    with open("script.bas", "w") as f:
        f.write(code)




#shutil.copy('name searched','file path distination')  copying
#scorce = 'name searched'
#destenation = ''
#filepath = filedialog.askopenfilename()
#file_name = os.path.basename(r'filepath')
# file name without extension
#hehe = (os.path.splitext(file_name)[0])

def choose_and_copy_srf():
    file_path = filedialog.askopenfilename(filetypes=[("SRF files", "*.srf")])

    if file_path:
        # Get the user's home directory
        home_dir = os.path.expanduser("~")

        # Create the destination folder if it doesn't exist
        destination_folder = os.path.join(home_dir, "Civigram")
        os.makedirs(destination_folder, exist_ok=True)

        # Check if there is an existing file with the name "sample.srf" in the destination folder
        destination_file = os.path.join(destination_folder, "sample.srf")
        if os.path.exists(destination_file):
            os.remove(destination_file)

        try:
            # Copy the surfer file and the script file to the destination folder
            file_name = os.path.basename(file_path)
            shutil.copyfile(file_path, destination_file)
            shutil.copy("script.bas", destination_folder)
            messagebox.showinfo("Info", f"Files {file_name} was uploaded successfully.")
        except Exception as e:
            messagebox.showwarning("Warning", f"An error occurred: {str(e)}")
    else:
        messagebox.showwarning("Warning", "No file selected.")

home_dir = os.path.expanduser("~")

# Create the destination folder if it doesn't exist
destination_folder = os.path.join(home_dir, "Civigram")


def texted():
    filepather = filedialog.askopenfilenames()
    for h in filepather:
        filepath = h
        file_name = os.path.basename(p=filepath)
        file_name_pure = (os.path.splitext(file_name)[0])
        df = pd.read_csv(fr'{filepath}')
        df5 = df.drop(df.columns[[0, -1]], axis=1)
        df1 = df5.assign(colname=(df5["N"]) * 10)
        cols = list(df1.columns)
        cols = [cols[-1]] + cols[:-1]
        df6 = df1[cols]
        df7 = df6.drop('N', axis=1)
        n = len(df7.columns)
        output = []
        for i in range(1, n):
            output.append(-(i - 1) * 2.5)
        output.insert(0, 0)
        df7.loc[len(df7)] = output
        df8 = df7.apply(np.roll, shift=1)
        df3 = pd.DataFrame(df8)
        products_list = df3.values.tolist()

        A = products_list
        l = []
        for i, y in enumerate(A[0][1:], 1):
            for z in A[1:]:
                x = [(z[0], y, z[i])]
                l.append(x)

        with open(f'{file_name_pure}.txt', 'w') as f:
            for line in l:
                text = f"{line}\n"
                patn = re.sub(r"[\([{}),\]]", "", text)
                patn = patn.replace(" ", "     ")
                f.write(patn)
        continue
    safe_as = filedialog.askdirectory()
    for g in filepather:
        filepath = g
        file_name = os.path.basename(p=filepath)
        file_name_pure = (os.path.splitext(file_name)[0])
        shutil.copy(fr'{file_name_pure}.txt', fr'{safe_as}')
        continue

def openfile():
    filepather = filedialog.askopenfilenames()
    for h in filepather:
        filepath = h
        file_name = os.path.basename(p=filepath)
        file_name_pure = (os.path.splitext(file_name)[0])
        df = pd.read_csv(fr'{filepath}')
        df1 = df.drop(df.columns[[0, 1, -1]], axis=1)
        df1.to_excel(fr'{file_name_pure}.xlsx')


        filename = fr'{file_name_pure}'
        wb1 = xl.load_workbook(fr'{filename}.xlsx')
        ws1 = wb1.worksheets[0]

        filename1 = "main.xlsx"
        wb2 = xl.load_workbook(filename1)
        ws2 = wb2.active

        mr = ws1.max_row
        mc = ws1.max_column

        for i in range(1, mr + 1):
            for j in range(1, mc + 1):
                c = ws1.cell(row=i + 1, column=j + 1)

                ws2.cell(row=i, column=j).value = c.value

        wb2.save(fr'{file_name_pure}.xlsx')




        continue


        # Define the folder where the files are located
        #current_dir = Path(__file__).parent if "__file__" in locals() else Path.cwd()
        #input_dir = current_dir / "INPUT"
        #output_dir = current_dir / "OUTPUT"
        #output_dir.mkdir(parents=True, exist_ok=True)

        # List all the excel files in the folder
        #xl_files = list(input_dir.rglob("*.xls*"))

        # List all the sheets that should be converted


    #for g in filepather:
        #filepath = g



        #continue








    safe_as = filedialog.askdirectory()
    for g in filepather:
        filepath = g
        file_name = os.path.basename(p=filepath)
        file_name_pure = (os.path.splitext(file_name)[0])
        shutil.copy(fr'{file_name_pure}', fr'{safe_as}')
        continue

def autocadedd():
    the_horizontal_drop = entry.get()
    the_vertical_drop = entry1.get()
    the_searched_freq = entry2.get()

    filepather = filedialog.askopenfilenames()
    for h in filepather:
        filepath = h
        file_name = os.path.basename(p=filepath)
        file_name_pure = (os.path.splitext(file_name)[0])
        df = pd.read_csv(fr'{filepath}')
        df5 = df.drop(df.columns[[0, -1]], axis=1)
        df1 = df5.assign(colname=(df5["N"]) * the_horizontal_drop)
        cols = list(df1.columns)
        cols = [cols[-1]] + cols[:-1]
        df6 = df1[cols]
        df7 = df6.drop('N', axis=1)
        n = len(df7.columns)
        output = []
        for i in range(1, n):
            output.append(-(i - 1) * the_vertical_drop)
        output.insert(0, 0)
        df7.loc[len(df7)] = output
        df8 = df7.apply(np.roll, shift=1)
        df3 = pd.DataFrame(df8)
        df3 = df3.drop('colname', axis=1)
        df3 = df3.drop(0)

        def search_row(value, row):
            if value <= row[0]:
                return 0
            elif value >= max(row):
                return len(row)
            else:
                for i in range(1, len(row)):
                    if row[i] >= value:
                        return i + (value - row[i - 1]) / (row[i] - row[i - 1])
                return len(row)

        value = the_searched_freq

        df3['search_result'] = df3.apply(lambda row: search_row(value, row), axis=1)

        df3['search_result'] = df3['search_result'].apply(lambda x: max(x - 1, 0) if x > 0 else x)

        search_result_list = df3['search_result'].tolist()

        acad = win32com.client.Dispatch("AutoCAD.Application")

        doc = acad.ActiveDocument

        now = datetime.datetime.now()
        ss_name = f"MySelectionSet_{now.strftime('%Y%m%d_%H%M%S')}"

        ss = doc.SelectionSets.Add(ss_name)
        ss.SelectOnScreen()

        points = []

        for i in range(ss.Count):
            entity = ss.Item(i)
            if entity.ObjectName == "AcDbPoint":  # Check if the entity is a point
                point = entity
                x = point.Coordinates[0]
                y = point.Coordinates[1]
                z = search_result_list[i]  # Replace the z coordinate with the value from the list
                points.append((x, y, z))  # Append the coordinates to the list

        line = None
        for entity in doc.ModelSpace:
            if entity.ObjectName == "AcDbLine":
                start_point = entity.StartPoint
                end_point = entity.EndPoint
                if all(point in [start_point, end_point] for point in points):
                    line = entity
                    break

        if line is not None:
            start_point = line.StartPoint
            end_point = line.EndPoint
            points.sort(key=lambda p: ((p[0] - start_point[0]) ** 2 + (p[1] - start_point[1]) ** 2 + (
                    p[2] - start_point[2]) ** 2) ** 0.5)
            if end_point != points[-1]:
                points.reverse()

        ss.Delete()

        points = list(reversed(points))

        from tkinter import scrolledtext

        def display_points():
            # Create a new window
            window2 = Toplevel()
            window2.title("Points")
            window2.geometry("400x400")

            # Create a Text widget to display the points
            points_text = scrolledtext.ScrolledText(window2, width=40, height=20)

            # Insert the points into the Text widget
            for i, point in enumerate(points):
                points_text.insert(END, f"{point[0]},{point[1]},{search_result_list[i] * the_vertical_drop * -1}\n")

            # Disable editing and enable copying of the points
            points_text.config(state=DISABLED)
            points_text.bind("<Control-a>", lambda event: points_text.tag_add(SEL, "1.0", END))
            points_text.bind("<Control-c>", lambda event: points_text.event_generate("<<Copy>>"))

            points_text.pack()

        # Create a button to display the points
        button = Button(window, text="Display Points", command=display_points)
        button.pack()

        window.mainloop()




#def copying_data():
    #df = pd.read_csv()
    #df1 = df.drop(df.columns[[0,1,-1]],axis=1)
    #df.to_excel('50M_L193.xlsx')



        #saving()





#min_column = wb.active.min_column
#max_column = wb.active.max_column
#min_row = wb.active.min_row
#max_row = wb.active.max_row





# Create the GUI
window = Tk()
window.title("'Data Manager by CIVIGRAM®")
window.config(background="#ffff00")
window.resizable(0, 0)  # Disable window resizing

font_style = ("Montserrat", 12)

lable = Label(window, text=" WE MAKE THINGS EASIER!!", font=('Montserrat', 15, 'bold'), fg='#000090',
              bg='#ffff00', padx=20, pady=20)
lable.pack()

lable10 = Label(window, text="Uplaod a new refrance(srf file): ", font=font_style, fg='#000090', bg='#ffff00',
               borderwidth=0)
lable10.pack()

button10 = Button(window, text="Upload", font=("Montserrat", 10), fg="black", bg="white", activebackground="#ffff00",
                 activeforeground="#000090", borderwidth=3, command=choose_and_copy_srf )
button10.config(font=font_style)
button10.pack()

lable1 = Label(window, text="The distance between the points(m): ", font=font_style, fg='#000090', bg='#ffff00',
               borderwidth=0)
lable1.pack()
def validate_entry(value):
    # Check if the value is numeric
    try:
        float(value)
        return True
    except ValueError:
        return False
entry = Entry(window, font=font_style, fg="#000090", bg="white", width=32, insertbackground="white", validate="key", validatecommand=(window.register(validate_entry), '%P'))
entry.pack()
entry.config(width=12)
entry.config(exportselection=False)

lable2 = Label(window, text="The drop between the readings(-m): ", font=font_style, fg='#000090', bg='#ffff00',
               borderwidth=0)
lable2.pack()
entry1 = Entry(window, font=font_style, fg="#000090", bg="white", width=32, insertbackground="white", validate="key", validatecommand=(window.register(validate_entry), '%P'))
entry1.pack()
entry1.config(width=12)
entry1.config(exportselection=False)

lable3 = Label(window, text="The frequency to search for: ", font=font_style, fg='#000090', bg='#ffff00',
               borderwidth=0)
lable3.pack()
entry2 = Entry(window, font=font_style, fg="#000090", bg="white", width=32, insertbackground="white",  validate="key", validatecommand=(window.register(validate_entry), '%P'))
entry2.pack()
entry2.config(width=12)
entry2.config(exportselection=False)

lable4 = Label(window, text="          ", bg='#ffff00').pack()

check_button = Checkbutton(window, text="Save the data as text '.txt'", onvalue=1, offvalue=0, font=font_style,
                           fg='#000090', bg='#ffff00', activeforeground='#000090',
                           activebackground='#ffff00', pady=3, padx=10, anchor='w')
#check_button.pack(anchor='w')

check_button1 = Checkbutton(window, text="Get the first appearance for a curtain frequency ", onvalue=1, offvalue=0,
                            font=font_style, fg='#000090', bg='#ffff00', activeforeground='#000090',
                            activebackground='#ffff00', pady=3, padx=10, anchor='w')
#check_button1.pack(anchor='w')

check_button2 = Checkbutton(window, text="Make a mesh and draw the profile", onvalue=1, offvalue=0, font=font_style,
                            fg='#000090', bg='#ffff00', activeforeground='#000090', activebackground='#ffff00',
                            pady=3, padx=10, anchor='w')
#check_button2.pack(anchor='w')


def autocaded():
    def error_handelling():
        the_horizontal_step = entry.get()
        the_vertical_drop = entry1.get()
        the_searched_freq = entry2.get()

        # Check if any of the values are empty
        if not the_horizontal_step or not the_vertical_drop or not the_searched_freq:
            # Display a warning message if any of the values are empty
            messagebox.showwarning("Empty input",
                                   "Please make sure that 'horizontal step, vertical drop, and searched frequency' aren't empty.")
            return
    error_handelling()

    def saving():
        # Get the selected file names from the listbox
        selected_file_names = [listbox.get(idx) for idx in listbox.curselection()]
        if not selected_file_names:
            # Display a warning message if no files are selected
            messagebox.showwarning("No files selected", "Please select one or more files.")
            return
        the_horizontal_step = entry.get()
        the_vertical_drop = entry1.get()
        the_searched_freq = entry2.get()

        # Check if any of the values are empty
        if not the_horizontal_step or not the_vertical_drop or not the_searched_freq:
            # Display a warning message if any of the values are empty
            messagebox.showwarning("Empty input",
                                   "Please make sure that 'horizontal step, vertical drop, and searched frequency' aren't empty.")
            return
        selected_file_names = [listbox.get(idx) for idx in listbox.curselection()]
        selected_file_paths = []
        selected_file_paths1 = []
        selected_file_paths2 = []
        for file_info in file_list:
            if file_info[2] in selected_file_names:
                selected_file_paths.append(file_info[0])
                selected_file_paths1.append(file_info[1])
                selected_file_paths2.append(file_info[2])
        safe_as = filedialog.askdirectory()
        for g in selected_file_paths:
            filepath = g
            file_name = os.path.basename(p=filepath)
            file_name_pure = (os.path.splitext(file_name)[0])
            new_file_path = os.path.join(safe_as, f"{file_name_pure}_done.srf")

            try:
                # Move the file to the new location
                shutil.move(os.path.join("C:\\Users\\", user_folder_name, "Civigram", f"{file_name_pure}_done.srf"),
                            new_file_path)
            except shutil.Error:
                # If the file already exists, ask the user if they want to replace it or save as a new file
                response = messagebox.askyesno("File already exists",
                                               f"The file '{file_name_pure}_done.srf' already exists in the selected directory. Do you want to replace it?")
                if response == True:
                    # If the user clicks "Replace", replace the existing file
                    shutil.move(os.path.join("C:\\Users\\", user_folder_name, "Civigram", f"{file_name_pure}_done.srf"),
                                new_file_path)
                else:
                    # If the user clicks "Save as", prompt the user to choose a new name and location for the file
                    new_file_path = filedialog.asksaveasfilename(initialdir=safe_as,
                                                                 initialfile=f"{file_name_pure}_done.srf")
                    if new_file_path:
                        # If the user chooses a new name and location, move the file to the new location
                        shutil.move(
                            os.path.join("C:\\Users\\", user_folder_name, "Civigram", f"{file_name_pure}_done.srf"),
                            new_file_path)

            for h in selected_file_paths:
                filepath = h
                file_name = os.path.basename(p=filepath)
                file_name_pure = (os.path.splitext(file_name)[0])
                os.remove(os.path.join("C:\\Users\\", user_folder_name, "Civigram", f"{file_name_pure}.txt"))
                #shutil.copy(os.path.join("C:\\Users\\", user_folder_name, "Civigram", f"{file_name_pure}_done.srf"),
                 #           new_file_path)

    def texted():
        # Get the selected file names from the listbox
        selected_file_names = [listbox.get(idx) for idx in listbox.curselection()]
        if not selected_file_names:
            # Display a warning message if no files are selected
            messagebox.showwarning("No files selected", "Please select one or more files.")
            return
        the_horizontal_step = entry.get()
        the_vertical_drop = entry1.get()
        the_searched_freq = entry2.get()

        # Check if any of the values are empty
        if not the_horizontal_step or not the_vertical_drop or not the_searched_freq:
            # Display a warning message if any of the values are empty
            messagebox.showwarning("Empty input",
                                   "Please make sure that 'horizontal step, vertical drop, and searched frequency' aren't empty.")
            return
        selected_file_names = [listbox.get(idx) for idx in listbox.curselection()]
        selected_file_paths = []
        selected_file_paths1 = []
        selected_file_paths2 = []
        for file_info in file_list:
            if file_info[2] in selected_file_names:
                selected_file_paths.append(file_info[0])
                selected_file_paths1.append(file_info[1])
                selected_file_paths2.append(file_info[2])

        # Get the values of the horizontal step, vertical drop, and searched frequency from the entry fields
        the_horizontal_step = entry.get()
        the_vertical_drop = entry1.get()
        the_searched_freq = entry2.get()

        # Create the destination folder if it doesn't exist
        destination_folder = os.path.join(os.path.expanduser("~"), "Civigram")
        os.makedirs(destination_folder, exist_ok=True)

        for h in selected_file_paths:
            filepath = h
            file_name = os.path.basename(p=filepath)
            file_name_pure = (os.path.splitext(file_name)[0])
            df = pd.read_csv(fr'{filepath}')
            df5 = df.drop(df.columns[[0, -1]], axis=1)
            df1 = df5.assign(colname=(df5["N"]) * float(the_horizontal_step))
            cols = list(df1.columns)
            cols = [cols[-1]] + cols[:-1]
            df6 = df1[cols]
            df7 = df6.drop('N', axis=1)
            n = len(df7.columns)
            output = []
            for i in range(1, n):
                output.append(-(i - 1) * float(the_vertical_drop))
            output.insert(0, 0)
            df7.loc[len(df7)] = output
            df8 = df7.apply(np.roll, shift=1)
            df3 = pd.DataFrame(df8)
            products_list = df3.values.tolist()

            A = products_list
            l = []
            for i, y in enumerate(A[0][1:], 1):
                for z in A[1:]:
                    x = [(z[0], y, z[i])]
                    l.append(x)

            with open(os.path.join(destination_folder, f'{file_name_pure}.txt'), 'w') as f:
                for line in l:
                    text = f"{line}\n"
                    patn = re.sub(r"[\([{}),\]]", "", text)
                    patn = patn.replace(" ", "     ")
                    f.write(patn)
            continue


        #def surfered():
            # Modify the path of the script file to include the default user and the file name
            #command = 'runas /user:defaultuser "C:\\Program Files\\Golden Software\\Surfer\\Scripter.exe" -x "C:\\Users\\defaultuser\\civigram\\script.bas"'

            # Launch the command as a subprocess without a window
            #startupinfo = subprocess.STARTUPINFO()
            #startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            #startupinfo.wShowWindow = subprocess.SW_HIDE  # hide the window
            #subprocess.Popen(command, startupinfo=startupinfo,
             #                creationflags=subprocess.CREATE_NO_WINDOW | subprocess.DETACHED_PROCESS)

        #surfered()
        def surfered():
            command = f'"C:\\Program Files\\Golden Software\\Surfer\\Scripter.exe" -x "C:\\Users\\{user_folder_name}\\civigram\\script.bas"'

            # Launch the command as a subprocess without a window
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow = subprocess.SW_HIDE  # hide the window
            subprocess.Popen(command, startupinfo=startupinfo, creationflags=subprocess.CREATE_NO_WINDOW | subprocess.DETACHED_PROCESS)
        surfered()




    # Define a function to update the Listbox with the file names
    def update_listbox():
        listbox.delete(0, END)  # Clear the existing items in the Listbox
        for file_info in file_list:
            listbox.insert(END, file_info[2])  # Add the file name to the Listbox

    # Create an empty list for the file paths and file names
    file_list = []

    # Define a function to retrieve the file paths based on the selected file names
    def get_file_paths_from_listbox():
        selected_file_names = [listbox.get(idx) for idx in listbox.curselection()]
        selected_file_paths = []
        for file_info in file_list:
            if file_info[2] in selected_file_names:
                selected_file_paths.append(file_info[0])
        the_horizontal_step = entry.get()
        the_vertical_drop = entry1.get()
        the_searched_freq = entry2.get()

        for h in selected_file_paths:
            df = pd.read_csv(fr'{h}')
            df5 = df.drop(df.columns[[0, -1]], axis=1)
            df1 = df5.assign(colname=(df5["N"]) * float(the_horizontal_step))
            cols = list(df1.columns)
            cols = [cols[-1]] + cols[:-1]
            df6 = df1[cols]
            df7 = df6.drop('N', axis=1)
            n = len(df7.columns)
            output = []
            for i in range(1, n):
                output.append(-(i - 1) * float(the_vertical_drop))
            output.insert(0, 0)
            df7.loc[len(df7)] = output
            df8 = df7.apply(np.roll, shift=1)
            df3 = pd.DataFrame(df8)
            df3 = df3.drop('colname', axis=1)
            df3 = df3.drop(0)

            def search_row(value, row):
                if value <= row[0]:
                    return 0
                elif value >= max(row):
                    return len(row)
                else:
                    for i in range(1, len(row)):
                        if row[i] >= value:
                            return i + (value - row[i - 1]) / (row[i] - row[i - 1])
                    return len(row)

            value = float(the_searched_freq)

            df3['search_result'] = df3.apply(lambda row: search_row(value, row), axis=1)

            df3['search_result'] = df3['search_result'].apply(lambda x: max(x - 1, 0) if x > 0 else x)

            search_result_list = df3['search_result'].tolist()

            # Connect to AutoCAD
            acad = win32com.client.Dispatch("AutoCAD.Application")

            # Get the active document
            doc = acad.ActiveDocument

            # Generate a unique selection set name using the current time
            now = datetime.datetime.now()
            ss_name = f"MySelectionSet_{now.strftime('%Y%m%d_%H%M%S')}"

            # Get the selection set of the active document
            ss = doc.SelectionSets.Add(ss_name)
            ss.SelectOnScreen()

            # Initialize a list to store the coordinates of each selected point
            points = []

            # Iterate over the selected entities
            for i in range(ss.Count):
                entity = ss.Item(i)
                if entity.ObjectName == "AcDbPoint":  # Check if the entity is a point
                    point = entity
                    x = point.Coordinates[0]
                    y = point.Coordinates[1]
                    z = point.Coordinates[2]
                    points.append((x, y, z))  # Append the coordinates to the list

            # Get the line that contains the selected points
            line = None
            for entity in doc.ModelSpace:
                if entity.ObjectName == "AcDbLine":
                    start_point = entity.StartPoint
                    end_point = entity.EndPoint
                    if all(point in [start_point, end_point] for point in points):
                        line = entity
                        break

            # Sort the list of points based on their position on the line
            if line is not None:
                start_point = line.StartPoint
                end_point = line.EndPoint
                points.sort(key=lambda p: ((p[0] - start_point[0]) ** 2 + (p[1] - start_point[1]) ** 2 + (
                        p[2] - start_point[2]) ** 2) ** 0.5)
                if end_point != points[-1]:
                    points.reverse()

            # Clear the selection set
            ss.Delete()

            # Reverse the order of the sorted points
            points = list(reversed(points))



            window = Tk()
            window.title("Points")
            window.geometry("400x400")

            # Create a Text widget to display the points
            points_text = scrolledtext.ScrolledText(window, width=80, height=40)

            # Insert the points into the Text widget
            for i, point in enumerate(points):
                points_text.insert(END,
                                   f"{point[0]},{point[1]},{search_result_list[i] * float(the_vertical_drop) * (-1)}\n")

            # Disable editing and enable copying of the points
            points_text.config(state=DISABLED)
            points_text.bind("<Control-a>", lambda event: points_text.tag_add(SEL, "1.0", END))
            points_text.bind("<Control-c>", lambda event: points_text.event_generate("<<Copy>>"))

            points_text.pack()

            window.mainloop()




    # Define a function to get the file paths and file names
    def get_file_paths():
        nonlocal file_list  # Use the 'nonlocal' keyword to modify the variable in the outer scope
        filepather = filedialog.askopenfilenames()
        for h in filepather:
            filepath = h
            file_name = os.path.basename(filepath)
            file_name_pure = os.path.splitext(file_name)[0]
            file_list.append((filepath, file_name, file_name_pure))
        update_listbox()  # Update the Listbox with the file names


    # Create a new window with a Listbox and two buttons
    new_window = Tk()
    new_window.title("the data")
    new_window.config(background="#ffff00")
    new_window.geometry("420x250")
    new_window.resizable(0, 0)

    listbox = Listbox(new_window, font=font_style, fg="#000090", bg="#F5F5DF", width=40, height=10, selectmode='multiple')
    listbox.pack()

    button1 = Button(new_window, text="Select Files", font=font_style, fg="black", bg="white",
                     activebackground="#ffff00",
                     activeforeground="#000090", command=get_file_paths)
    button1.pack(side='left', padx=10, pady=10)

    button2 = Button(new_window, text="draw cad", font=font_style, fg="black", bg="white",
                     activebackground="#ffff00",
                     activeforeground="#000090", command=get_file_paths_from_listbox)
    button2.pack(side='right', padx=10, pady=10)

    button3 = Button(new_window, text="save", font=font_style, fg="black", bg="white",
                     activebackground="#ffff00",
                     activeforeground="#000090", command=saving)
    button3.pack(side='right', padx=10, pady=10)


    button4 = Button(new_window, text="draw surfer", font=font_style, fg="black", bg="white",
                     activebackground="#ffff00",
                     activeforeground="#000090", command=texted)
    button4.pack(side='right', padx=10, pady=10)

    new_window.mainloop()


button1 = Button(window, text="Start", font=("Montserrat", 10), fg="black", bg="white", activebackground="#ffff00",
                 activeforeground="#000090", command=autocaded, borderwidth=3)
button1.config(font=font_style)
button1.pack()

window.mainloop()
#from openpyxl import load_workbook

#wb = load_workbook(fr'{file_name_pure}.xlsx')
#sheet = wb['CHART']

#min_column = wb.active.min_column
#max_column = wb.active.max_column
#min_row = wb.active.min_row
#max_row = wb.active.max_row
































